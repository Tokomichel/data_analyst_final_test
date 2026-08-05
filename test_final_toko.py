#!/usr/bin/env python3
"""Recherche d'infos sur une entreprise via l'API Perplexity."""

import json
import time
import httpx
from openai import OpenAI, APIConnectionError, AuthenticationError, APITimeoutError
import openpyxl
import dotenv

dotenv.load_dotenv()

# On recupere la clé de l'api depuis le fichier .env (à la racine du projet)
API_KEY = dotenv.get_key(dotenv.find_dotenv(), "API_KEY")

SYSTEM_PROMPT = """Tu es un assistant de recherche. Pour l'entreprise demandée, trouve:
1. Le nom du CEO actuel
2. Le lien LinkedIn du CEO
3. Le nombre d'employés
4. Le pays de résidence du CEO

Réponds UNIQUEMENT en JSON valide:
{
  "company": "nom",
  "ceo_name": "Nom complet",
  "ceo_linkedin": "https://linkedin.com/in/... ou null",
  "employee_count": "nombre ou fourchette",
  "ceo_residence_country": "Pays"
}
Si introuvable, mets null. Réponds en français."""


FIELDS = [
    "ceo_name",
    "ceo_linkedin",
    "employee_count",
    "ceo_residence_country",
]
 
HEADERS = ["Company Name"] + [f.replace("_", " ").title() for f in FIELDS]


def recherche_entreprise(
    api_key: str,
    nom_entreprise: str,
    model: str = "sonar",
    proxy: str = None,
    timeout: int = 60,
) -> dict:
    """
    Recherche le CEO, son LinkedIn, le nombre d'employés et le pays de résidence.

    Paramètres:
        api_key        : clé API Perplexity (commence par 'pplx-')
        nom_entreprise : nom de l'entreprise à rechercher
        model          : modèle à utiliser ("sonar" ou "sonar-pro")
        proxy          : URL de proxy si besoin (ex: "http://proxy:8080")
        timeout        : délai max en secondes (défaut: 60)

    Retourne:
        {
            "company": str,
            "ceo_name": str | None,
            "ceo_linkedin": str | None,
            "employee_count": str | None,
            "ceo_residence_country": str | None,
            "citations": list[str],
            "error": str | None
        }
    """
    result = {
        "company": nom_entreprise,
        "ceo_name": None,
        "ceo_linkedin": None,
        "employee_count": None,
        "ceo_residence_country": None,
        "citations": [],
        "error": None,
    }

    # ── Client avec vérification SSL désactivée ──
    # Contourne l'erreur Windows CRYPT_E_NO_REVOCATION_CHECK
    # (réseau d'entreprise, antivirus, proxy qui intercepte le HTTPS)
    http_client = httpx.Client(verify=False, proxy=proxy, timeout=timeout)

    client = OpenAI(
        api_key=API_KEY,
        base_url="https://api.perplexity.ai",
        http_client=http_client,
        timeout=timeout,
    )

    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": f"Entreprise : {nom_entreprise}"},
            ],
            temperature=0.2,
        )

        raw = resp.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1].rsplit("```", 1)[0].strip()

        data = json.loads(raw)
        result.update({
            "ceo_name": data.get("ceo_name"),
            "ceo_linkedin": data.get("ceo_linkedin"),
            "employee_count": data.get("employee_count"),
            "ceo_residence_country": data.get("ceo_residence_country"),
        })

        if hasattr(resp, "citations") and resp.citations:
            result["citations"] = resp.citations

    except APIConnectionError as e:
        result["error"] = f"Connexion impossible: {e}"
    except AuthenticationError:
        result["error"] = "Clé API invalide. Vérifie qu'elle commence par 'pplx-'."
    except APITimeoutError:
        result["error"] = f"Délai dépassé ({timeout}s). Augmente le paramètre timeout."
    except Exception as e:
        result["error"] = f"Erreur inattendue: {type(e).__name__}: {e}"

    return result

# ── Exemple d'utilisation ────────────────────────────────────────────
if __name__ == "__main__":
    import warnings
    warnings.filterwarnings("ignore")  # masque le warning du SSL désactivé

    wb = openpyxl.load_workbook("Test_final.xlsx")
    ws = wb.active
    

    # Écrit les en-têtes en ligne 1 
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    row_idx = 2

    while True: # Limite à 100 lignes qui est le nombre de lignes du fichier Excel

        company_cell = ws.cell(row=row_idx, column=1)
        company_name = company_cell.value
 
        # Arrête si la colonne A est vide (fin du tableau)
        if company_name is None or str(company_name).strip() == "":
            print("La colonne A est vide on s'arrete là")
            break
 
        print(f"Ligne {row_idx} : {company_name}")

        try:
            result = recherche_entreprise(API_KEY, company_name, timeout=5)
        except Exception as e:
            print(f"  -> Erreur API pour '{company_name}': {e}")
            continue

        print(f"\n{company_name} :")
        print(f"  CEO       : {result['ceo_name']}")
        print(f"  LinkedIn  : {result['ceo_linkedin']}")
        print(f"  Employés  : {result['employee_count']}")
        print(f"  Pays      : {result['ceo_residence_country']}")
        print(f"index : {row_idx}")
        if result["error"]:
            print(f"  Erreur    : {result['error']}")

        # Écrit chaque champ dans sa colonne (B=2, C=3, ...)
        for offset, field in enumerate(FIELDS, start=2):
            value = result.get(field, "")
            if value is None:
                value = "None"
            ws.cell(row=row_idx, column=offset, value=value)
        row_idx += 1

    wb.save("Test_final_result.xlsx")



